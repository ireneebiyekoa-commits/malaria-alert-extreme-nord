"""Mise à jour des données (admin uniquement)."""
import io
from calendar import monthrange
from datetime import date, datetime

import pandas as pd
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse

from apps.accounts.decorators import admin_required
from apps.core.models import (District, ImportLog, Meteo, Observation,
                              Prevision, SeuilAlerte)
from apps.predictions.engine import predict_recursive
from apps.predictions.ml_loader import is_ready

from .nasa_power import fetch_climatic_data


@admin_required
def index(request):
    """Page principale de mise à jour."""
    # Dernier import + dernière date observée
    derniere_obs = Observation.objects.order_by('-date').first()
    derniers_imports = ImportLog.objects.all()[:10]

    # Période suggérée : mois suivant la dernière observation
    if derniere_obs:
        next_date = (pd.Timestamp(derniere_obs.date) + pd.DateOffset(months=1))
        annee_suggeree = int(next_date.year)
        mois_suggere = int(next_date.month)
    else:
        annee_suggeree = date.today().year
        mois_suggere = date.today().month

    context = {
        'derniere_observation': derniere_obs.date if derniere_obs else None,
        'derniers_imports': derniers_imports,
        'annee_suggeree': annee_suggeree,
        'mois_suggere': mois_suggere,
        'nb_districts': District.objects.count(),
    }
    return render(request, 'data_management/index.html', context)


@admin_required
def telecharger_template(request):
    """Télécharge un template Excel pré-rempli pour la saisie mensuelle."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    annee = int(request.GET.get('annee', date.today().year))
    mois = int(request.GET.get('mois', date.today().month))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Saisie {annee}-{mois:02d}"

    # En-tête
    headers = ['district', 'annee', 'mois', 'cas_confirmes', 'population']
    header_fill = PatternFill(start_color="0E476E", end_color="0E476E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Pré-remplissage : district + population (dernière connue), mois précédent en commentaire
    districts = District.objects.all().order_by('nom')
    for row_idx, d in enumerate(districts, 2):
        ws.cell(row=row_idx, column=1, value=d.nom)
        ws.cell(row=row_idx, column=2, value=annee)
        ws.cell(row=row_idx, column=3, value=mois)
        ws.cell(row=row_idx, column=4, value=None)  # à remplir
        ws.cell(row=row_idx, column=5, value=d.population)

    for col_letter in 'ABCDE':
        ws.column_dimensions[col_letter].width = 22

    # Feuille d'instructions
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "INSTRUCTIONS DE SAISIE — Système d'alerte paludisme",
        "",
        "1. Renseigner uniquement la colonne 'cas_confirmes' pour chaque district.",
        "2. Ne pas modifier les colonnes district, année, mois.",
        "3. La colonne population peut être actualisée si nécessaire.",
        "4. Ne pas ajouter ni supprimer de lignes.",
        "5. Une fois terminé, importer ce fichier via la page 'Mise à jour des données'.",
        "",
        f"Période de saisie : {mois:02d}/{annee}",
        f"Districts à renseigner : {districts.count()}",
        "",
        "© MINSANTE / GTR Paludisme Extrême-Nord",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="0E476E")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="saisie_paludisme_{annee}-{mois:02d}.xlsx"'
    )
    return response


@admin_required
def importer_donnees(request):
    """Importe un fichier Excel + récupère le climat via NASA POWER + recalcule les prévisions."""
    if request.method != 'POST':
        return redirect('data_management:index')

    fichier = request.FILES.get('fichier_excel')
    if not fichier:
        messages.error(request, "Aucun fichier fourni.")
        return redirect('data_management:index')

    try:
        df = pd.read_excel(fichier)
    except Exception as exc:
        messages.error(request, f"Impossible de lire le fichier : {exc}")
        return redirect('data_management:index')

    required_cols = {'district', 'annee', 'mois', 'cas_confirmes', 'population'}
    if not required_cols.issubset(df.columns):
        missing = required_cols - set(df.columns)
        messages.error(request, f"Colonnes manquantes : {missing}")
        return redirect('data_management:index')

    # Validation : pas de valeurs manquantes critiques
    if df['cas_confirmes'].isna().any():
        messages.error(request, "Le fichier contient des cas_confirmes manquants.")
        return redirect('data_management:index')

    districts = {d.nom: d for d in District.objects.all()}
    nb_imported = 0
    nb_meteo = 0
    nb_errors = 0
    periodes = set()

    with transaction.atomic():
        for _, row in df.iterrows():
            nom = str(row['district']).strip()
            district = districts.get(nom)
            if not district:
                nb_errors += 1
                continue

            annee = int(row['annee'])
            mois = int(row['mois'])
            date_obs = date(annee, mois, 1)
            periodes.add(f"{annee}-{mois:02d}")

            # Observation
            Observation.objects.update_or_create(
                district=district,
                date=date_obs,
                defaults={
                    'annee': annee,
                    'mois': mois,
                    'trimestre': f"T{((mois - 1) // 3) + 1}",
                    'cas_confirmes': float(row['cas_confirmes']),
                    'population': int(row['population']),
                }
            )
            nb_imported += 1

            # Météo via NASA POWER
            meteo_data = fetch_climatic_data(district.longitude, district.latitude, annee, mois)
            if meteo_data:
                Meteo.objects.update_or_create(
                    district=district,
                    date=date_obs,
                    defaults={**meteo_data, 'source': 'NASA_POWER'},
                )
                nb_meteo += 1
            else:
                # Fallback climatologie
                from apps.predictions.ml_loader import get_climatologie
                clim = get_climatologie()
                if clim is not None:
                    crow = clim[(clim['district'] == nom) & (clim['mois'] == mois)]
                    if not crow.empty:
                        Meteo.objects.update_or_create(
                            district=district, date=date_obs,
                            defaults={
                                'temp_moy': float(crow['temp_clim'].iloc[0]),
                                'humidite': float(crow['hum_clim'].iloc[0]),
                                'precip_mensuel': float(crow['precip_clim'].iloc[0]),
                                'source': 'climatologie_fallback',
                            }
                        )

            # Mettre à jour la population du district
            district.population = int(row['population'])
            district.save(update_fields=['population'])

        # Journal
        ImportLog.objects.create(
            utilisateur=request.user,
            fichier_nom=fichier.name,
            nb_lignes=nb_imported,
            nb_districts=len({n for n in df['district'].unique()}),
            periode=', '.join(sorted(periodes)),
            statut='succes' if nb_errors == 0 else 'partiel',
            message=f"{nb_imported} observations / {nb_meteo} météo / {nb_errors} erreurs",
        )

    # Recalcul des prévisions
    nb_previsions = 0
    if is_ready():
        nb_previsions = recalculer_previsions()

    messages.success(
        request,
        f"Import réussi : {nb_imported} observations, {nb_meteo} météo NASA POWER, "
        f"{nb_previsions} prévisions recalculées."
    )
    if nb_errors:
        messages.warning(request, f"{nb_errors} ligne(s) ignorée(s) (district inconnu).")

    return redirect('data_management:index')


def recalculer_previsions():
    """Recalcule les prévisions pour tous les districts (32 x 3 x 2 = 192)."""
    nb = 0
    seuils_map = {(s.district_id, s.mois_calendaire): s
                  for s in SeuilAlerte.objects.all()}

    with transaction.atomic():
        for district in District.objects.all():
            obs = list(Observation.objects.filter(district=district).order_by('date'))
            meteo = {m.date: m for m in Meteo.objects.filter(district=district)}
            if not obs:
                continue

            rows = []
            for o in obs:
                m = meteo.get(o.date)
                rows.append({
                    'date': o.date, 'annee': o.annee, 'mois': o.mois,
                    'incidence': o.incidence,
                    'precip_mensuel': m.precip_mensuel if m else 0,
                    'temp_moy': m.temp_moy if m else 0,
                    'humidite': m.humidite if m else 0,
                    'population': o.population,
                })
            historic_df = pd.DataFrame(rows)
            derniere_date = obs[-1].date

            for algo in ['RF', 'XGB']:
                try:
                    preds = predict_recursive(algo, district.nom, historic_df, horizons=[1, 2, 3])
                except Exception:
                    continue

                for h, p in preds.items():
                    seuil = seuils_map.get((district.id, p['mois_cible']))
                    s_alerte = seuil.seuil_alerte if seuil else 0
                    s_epidemio = seuil.seuil_epidemio if seuil else 0
                    niveau = 'vert'
                    if seuil:
                        if p['incidence'] >= seuil.seuil_epidemio:
                            niveau = 'rouge'
                        elif p['incidence'] >= seuil.seuil_alerte:
                            niveau = 'orange'
                    cas = (p['incidence'] * district.population) / 1000

                    Prevision.objects.update_or_create(
                        district=district,
                        algorithme=algo,
                        horizon=h,
                        date_origine=derniere_date,
                        defaults={
                            'date_cible': p['date_cible'],
                            'incidence_predite': p['incidence'],
                            'cas_predits': cas,
                            'niveau_alerte': niveau,
                            # Conservé pour compat schéma (les valeurs sont équivalentes)
                            'seuil_p75': s_alerte,
                            'seuil_p90': s_epidemio,
                        }
                    )
                    nb += 1
    return nb
