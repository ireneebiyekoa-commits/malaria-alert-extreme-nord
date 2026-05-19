"""Vues — page Prévisions."""
import json
from datetime import date

import pandas as pd
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_POST

from apps.accounts.decorators import admin_required
from apps.core.models import District, Meteo, Observation, Performance, SeuilAlerte
from apps.core.utils import filter_by_user_role, format_mois_annee

from .ai_analyzer import analyser_previsions
from .chatbot import repondre as chatbot_repondre
from .engine import predict_recursive
from .ml_loader import get_status, is_ready
from .report_generator import generer_rapport_previsions


@login_required
def index(request):
    """Page Prévisions — interface de sélection."""
    user = request.user

    if user.is_admin:
        districts = District.objects.all().order_by('nom')
    elif user.district_id:
        districts = District.objects.filter(pk=user.district_id)
    else:
        districts = District.objects.none()

    context = {
        'districts': districts,
        'algorithmes': [('RF', 'Random Forest'), ('XGB', 'XGBoost'), ('COMP', 'Comparaison RF / XGB')],
        'horizons': [1, 2, 3],
        'ml_status': get_status(),
        'ml_ready': is_ready(),
    }
    return render(request, 'predictions/index.html', context)


@login_required
def api_prevision(request):
    """API : retourne les prévisions pour un district, un algo et un horizon donnés."""
    user = request.user
    district_id = request.GET.get('district_id')
    algo = request.GET.get('algo', 'XGB').upper()
    horizon = int(request.GET.get('horizon', 3))

    if not district_id:
        return JsonResponse({'error': 'district_id requis'}, status=400)

    district = get_object_or_404(District, pk=district_id)
    if not user.peut_voir_district(district):
        return JsonResponse({'error': 'accès refusé'}, status=403)

    if not is_ready():
        return JsonResponse({
            'error': 'Modèles ML non chargés',
            'status': get_status(),
        }, status=503)

    # Construire l'historique du district
    obs = list(Observation.objects.filter(district=district).order_by('date'))
    meteo = {m.date: m for m in Meteo.objects.filter(district=district)}

    if not obs:
        return JsonResponse({'error': 'Aucune observation pour ce district'}, status=404)

    rows = []
    for o in obs:
        m = meteo.get(o.date)
        rows.append({
            'date': o.date,
            'annee': o.annee,
            'mois': o.mois,
            'incidence': o.incidence,
            'precip_mensuel': m.precip_mensuel if m else 0,
            'temp_moy': m.temp_moy if m else 0,
            'humidite': m.humidite if m else 0,
            'population': o.population,
        })
    historic_df = pd.DataFrame(rows)

    # Seuils d'alerte
    seuils = {s.mois_calendaire: s for s in SeuilAlerte.objects.filter(district=district)}

    horizons = list(range(1, horizon + 1))

    def _build_predictions(algo_code):
        preds = predict_recursive(algo_code, district.nom, historic_df, horizons=horizons)
        result = []
        for h, p in preds.items():
            seuil = seuils.get(p['mois_cible'])
            niveau = 'vert'
            p75 = seuil.p75 if seuil else 0
            p90 = seuil.p90 if seuil else 0
            if seuil:
                if p['incidence'] >= seuil.p90:
                    niveau = 'rouge'
                elif p['incidence'] >= seuil.p75:
                    niveau = 'orange'
            cas = (p['incidence'] * district.population) / 1000
            result.append({
                'horizon': h,
                'date': p['date_cible'].isoformat(),
                'mois_label': format_mois_annee(p['date_cible']),
                'incidence': round(p['incidence'], 3),
                'cas': round(cas, 0),
                'niveau': niveau,
                'p75': round(p75, 2),
                'p90': round(p90, 2),
            })
        return result

    if algo == 'COMP':
        previsions_rf = _build_predictions('RF')
        previsions_xgb = _build_predictions('XGB')
        previsions = {'RF': previsions_rf, 'XGB': previsions_xgb}
    else:
        previsions = _build_predictions(algo)

    # Métriques agrégées sur les 5 folds
    perfs = Performance.objects.filter(algorithme=algo if algo != 'COMP' else 'XGB',
                                       horizon=horizon)
    if perfs.exists():
        metriques = {
            'rmse': round(perfs.aggregate(m=Avg('rmse'))['m'] or 0, 3),
            'mae': round(perfs.aggregate(m=Avg('mae'))['m'] or 0, 3),
            'r2': round(perfs.aggregate(m=Avg('r2'))['m'] or 0, 3),
        }
    else:
        metriques = {'rmse': 0, 'mae': 0, 'r2': 0}

    # Historique récent pour le graphique
    derniere = obs[-1].date
    historique_12m = []
    for o in obs[-12:]:
        historique_12m.append({
            'date': o.date.isoformat(),
            'mois_label': format_mois_annee(o.date),
            'incidence': round(o.incidence, 3),
            'cas': o.cas_confirmes,
        })

    return JsonResponse({
        'district': district.nom,
        'district_court': district.nom_court,
        'population': district.population,
        'algorithme': algo,
        'horizon': horizon,
        'previsions': previsions,
        'historique': historique_12m,
        'metriques': metriques,
        'derniere_observation': derniere.isoformat(),
    })


@login_required
@require_POST
def api_chat(request):
    """Endpoint du chatbot analytique (strictement périmétré)."""
    payload = json.loads(request.body or '{}')
    question = payload.get('question', '')
    historique = payload.get('historique', [])
    contexte_district = payload.get('district')

    result = chatbot_repondre(
        question=question,
        historique=historique,
        contexte_district=contexte_district,
    )
    return JsonResponse(result)


@login_required
@require_POST
def api_analyse_ia(request):
    """Génère une analyse IA des prévisions affichées."""
    payload = json.loads(request.body or '{}')

    district = payload.get('district', '')
    algorithme = payload.get('algorithme', 'XGBoost')
    previsions = payload.get('previsions', [])
    historique = payload.get('historique', [])
    metriques = payload.get('metriques', {})

    if not previsions:
        return JsonResponse({'success': False, 'erreur': 'Aucune prévision fournie'}, status=400)

    result = analyser_previsions(
        district=district,
        algorithme=algorithme,
        previsions=previsions,
        historique_recent=historique,
        metriques=metriques,
    )
    return JsonResponse(result)


@login_required
@admin_required
@require_POST
def api_generer_rapport(request):
    """Génère et télécharge un rapport Word des prévisions (admin only)."""
    payload = json.loads(request.body or '{}')

    district = payload.get('district', '—')
    algorithme = payload.get('algorithme', 'XGBoost')
    horizon = int(payload.get('horizon', 3))
    previsions = payload.get('previsions', [])
    historique = payload.get('historique', [])
    metriques = payload.get('metriques', {})
    analyse_ia = payload.get('analyse_ia', '')

    if not previsions:
        return JsonResponse({'error': 'Aucune prévision fournie'}, status=400)

    buf = generer_rapport_previsions(
        district=district,
        algorithme=algorithme,
        horizon=horizon,
        previsions=previsions,
        historique=historique,
        metriques=metriques,
        analyse_ia=analyse_ia or "Analyse non générée.",
        auteur=request.user.get_full_name() or request.user.username,
    )

    filename = f"rapport_prevision_{district.replace(' ', '_')}_{date.today().isoformat()}.docx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def api_export_excel_global(request):
    """
    Export Excel global : toutes les prévisions (32 districts × RF + XGB × h=1,2,3)
    dans un seul fichier multi-feuilles.

    Feuilles :
      - Synthèse        : un tableau condensé (district x horizon)
      - Prévisions RF   : détail RF
      - Prévisions XGB  : détail XGB
      - Comparaison     : RF vs XGB côte à côte
      - Récap alertes   : niveaux par district et horizon
      - Performances    : métriques walk-forward
    """
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    user = request.user
    if not is_ready():
        return JsonResponse({'error': 'Modèles ML non chargés'}, status=503)

    # Filtrage par rôle
    if user.is_admin:
        districts_qs = District.objects.all().order_by('nom')
    elif user.district_id:
        districts_qs = District.objects.filter(pk=user.district_id)
    else:
        districts_qs = District.objects.none()

    # ---- Calcul de toutes les prévisions ----
    all_results = []
    seuils_map = {(s.district_id, s.mois_calendaire): s for s in SeuilAlerte.objects.all()}

    for district in districts_qs:
        obs = list(Observation.objects.filter(district=district).order_by('date'))
        meteos = {m.date: m for m in Meteo.objects.filter(district=district)}
        if not obs:
            continue

        rows = []
        for o in obs:
            m = meteos.get(o.date)
            rows.append({
                'date': o.date, 'annee': o.annee, 'mois': o.mois,
                'incidence': o.incidence,
                'precip_mensuel': m.precip_mensuel if m else 0,
                'temp_moy': m.temp_moy if m else 0,
                'humidite': m.humidite if m else 0,
                'population': o.population,
            })
        historic_df = pd.DataFrame(rows)

        for algo in ['RF', 'XGB']:
            try:
                preds = predict_recursive(algo, district.nom, historic_df, horizons=[1, 2, 3])
            except Exception:
                continue
            for h, p in preds.items():
                seuil = seuils_map.get((district.id, p['mois_cible']))
                p75 = seuil.p75 if seuil else 0
                p90 = seuil.p90 if seuil else 0
                niveau = 'vert'
                if seuil:
                    if p['incidence'] >= seuil.p90:
                        niveau = 'rouge'
                    elif p['incidence'] >= seuil.p75:
                        niveau = 'orange'
                cas = (p['incidence'] * district.population) / 1000
                all_results.append({
                    'district': district.nom,
                    'district_court': district.nom_court,
                    'algo': algo,
                    'horizon': h,
                    'date_cible': p['date_cible'],
                    'mois_label': format_mois_annee(p['date_cible']),
                    'incidence': round(p['incidence'], 3),
                    'cas': round(cas, 0),
                    'p75': round(p75, 2),
                    'p90': round(p90, 2),
                    'niveau': niveau,
                    'population': district.population,
                })

    if not all_results:
        return JsonResponse({'error': 'Aucune prévision générée'}, status=500)

    # ============================================================
    # Construction du workbook Excel multi-feuilles
    # ============================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # supprimer la feuille par défaut

    # Styles
    header_fill = PatternFill(start_color="0E476E", end_color="0E476E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="0E476E", size=14)
    sub_font = Font(italic=True, color="5A6878", size=10)
    border = Border(*[Side(border_style='thin', color='CCCCCC') for _ in range(4)])
    alert_fill = {
        'vert':   PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'orange': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'rouge':  PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }

    def _apply_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

    def _autosize(ws, widths=None):
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        else:
            for col_cells in ws.columns:
                length = max((len(str(c.value or '')) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    today = date.today().strftime('%d/%m/%Y')

    # ---------- Feuille 0 : Page de garde ----------
    ws0 = wb.create_sheet("Synthèse")
    ws0.merge_cells('A1:G1')
    ws0['A1'] = "EXPORT GLOBAL DES PRÉVISIONS ÉPIDÉMIQUES"
    ws0['A1'].font = title_font
    ws0['A1'].alignment = Alignment(horizontal='center')
    ws0.merge_cells('A2:G2')
    ws0['A2'] = f"Région de l'Extrême-Nord du Cameroun — Généré le {today}"
    ws0['A2'].font = sub_font
    ws0['A2'].alignment = Alignment(horizontal='center')

    # Tableau condensé : un district par ligne, colonnes horizon 1/2/3 pour chaque algo
    headers = ['Rang', 'District', 'Population',
               'RF h=1', 'RF h=2', 'RF h=3',
               'XGB h=1', 'XGB h=2', 'XGB h=3',
               'Niveau RF h=1', 'Niveau XGB h=1']
    _apply_header(ws0, headers, row=4)

    # Agrégation par district
    par_district = {}
    for r in all_results:
        key = r['district']
        if key not in par_district:
            par_district[key] = {'district_court': r['district_court'],
                                 'population': r['population']}
        par_district[key][f"{r['algo']}_{r['horizon']}"] = r['incidence']
        par_district[key][f"{r['algo']}_{r['horizon']}_niveau"] = r['niveau']

    # Trier par incidence XGB h=1 décroissante
    sorted_d = sorted(par_district.items(),
                      key=lambda kv: kv[1].get('XGB_1', 0), reverse=True)

    for idx, (nom, d) in enumerate(sorted_d, 1):
        row_num = 4 + idx
        ws0.cell(row=row_num, column=1, value=idx)
        ws0.cell(row=row_num, column=2, value=d['district_court'])
        ws0.cell(row=row_num, column=3, value=d['population'])
        ws0.cell(row=row_num, column=4, value=round(d.get('RF_1', 0), 2))
        ws0.cell(row=row_num, column=5, value=round(d.get('RF_2', 0), 2))
        ws0.cell(row=row_num, column=6, value=round(d.get('RF_3', 0), 2))
        ws0.cell(row=row_num, column=7, value=round(d.get('XGB_1', 0), 2))
        ws0.cell(row=row_num, column=8, value=round(d.get('XGB_2', 0), 2))
        ws0.cell(row=row_num, column=9, value=round(d.get('XGB_3', 0), 2))
        niv_rf = d.get('RF_1_niveau', 'vert')
        niv_xgb = d.get('XGB_1_niveau', 'vert')
        c_rf = ws0.cell(row=row_num, column=10, value=niv_rf.upper())
        c_rf.fill = alert_fill.get(niv_rf, alert_fill['vert'])
        c_rf.alignment = Alignment(horizontal='center')
        c_xgb = ws0.cell(row=row_num, column=11, value=niv_xgb.upper())
        c_xgb.fill = alert_fill.get(niv_xgb, alert_fill['vert'])
        c_xgb.alignment = Alignment(horizontal='center')

    _autosize(ws0, [6, 22, 14, 10, 10, 10, 11, 11, 11, 16, 16])
    ws0.freeze_panes = 'A5'

    # ---------- Feuilles : Détail RF + Détail XGB ----------
    for algo in ['RF', 'XGB']:
        ws = wb.create_sheet(f"Prévisions {algo}")
        _apply_header(ws, [
            'District', 'Horizon', 'Date cible',
            'Incidence prédite (/1000)', 'Cas attendus',
            'Seuil P75', 'Seuil P90', "Niveau d'alerte", 'Population'
        ])
        algo_results = [r for r in all_results if r['algo'] == algo]
        for idx, r in enumerate(algo_results, 2):
            ws.cell(row=idx, column=1, value=r['district_court'])
            ws.cell(row=idx, column=2, value=f"h={r['horizon']}")
            ws.cell(row=idx, column=3, value=r['mois_label'])
            ws.cell(row=idx, column=4, value=r['incidence'])
            ws.cell(row=idx, column=5, value=int(r['cas']))
            ws.cell(row=idx, column=6, value=r['p75'])
            ws.cell(row=idx, column=7, value=r['p90'])
            c_niv = ws.cell(row=idx, column=8, value=r['niveau'].upper())
            c_niv.fill = alert_fill.get(r['niveau'], alert_fill['vert'])
            c_niv.alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=9, value=r['population'])
        _autosize(ws, [20, 10, 18, 22, 14, 12, 12, 16, 14])
        ws.freeze_panes = 'A2'

    # ---------- Feuille : Comparaison RF vs XGB ----------
    ws_comp = wb.create_sheet("Comparaison RF vs XGB")
    _apply_header(ws_comp, [
        'District', 'Horizon', 'Date cible',
        'RF (inc.)', 'XGB (inc.)', 'Écart absolu', 'Écart relatif (%)',
        'Niveau RF', 'Niveau XGB', 'Concordance'
    ])
    # Joindre RF et XGB par (district, horizon)
    rf_map = {(r['district'], r['horizon']): r for r in all_results if r['algo'] == 'RF'}
    xgb_map = {(r['district'], r['horizon']): r for r in all_results if r['algo'] == 'XGB'}
    idx_comp = 2
    for key in sorted(rf_map.keys(), key=lambda k: (k[0], k[1])):
        rf_r = rf_map[key]
        xgb_r = xgb_map.get(key)
        if not xgb_r:
            continue
        ecart = round(abs(rf_r['incidence'] - xgb_r['incidence']), 3)
        rel = round((ecart / xgb_r['incidence'] * 100) if xgb_r['incidence'] else 0, 1)
        concordance = "✓ Identique" if rf_r['niveau'] == xgb_r['niveau'] else "✗ Divergent"
        ws_comp.cell(row=idx_comp, column=1, value=rf_r['district_court'])
        ws_comp.cell(row=idx_comp, column=2, value=f"h={rf_r['horizon']}")
        ws_comp.cell(row=idx_comp, column=3, value=rf_r['mois_label'])
        ws_comp.cell(row=idx_comp, column=4, value=rf_r['incidence'])
        ws_comp.cell(row=idx_comp, column=5, value=xgb_r['incidence'])
        ws_comp.cell(row=idx_comp, column=6, value=ecart)
        ws_comp.cell(row=idx_comp, column=7, value=rel)
        c1 = ws_comp.cell(row=idx_comp, column=8, value=rf_r['niveau'].upper())
        c1.fill = alert_fill.get(rf_r['niveau'], alert_fill['vert'])
        c1.alignment = Alignment(horizontal='center')
        c2 = ws_comp.cell(row=idx_comp, column=9, value=xgb_r['niveau'].upper())
        c2.fill = alert_fill.get(xgb_r['niveau'], alert_fill['vert'])
        c2.alignment = Alignment(horizontal='center')
        ws_comp.cell(row=idx_comp, column=10, value=concordance)
        idx_comp += 1
    _autosize(ws_comp, [20, 10, 18, 11, 11, 14, 18, 12, 12, 16])
    ws_comp.freeze_panes = 'A2'

    # ---------- Feuille : Récap alertes ----------
    ws_alerts = wb.create_sheet("Récap alertes")
    _apply_header(ws_alerts, ['District', 'Niveau RF h=1', 'Niveau RF h=2', 'Niveau RF h=3',
                              'Niveau XGB h=1', 'Niveau XGB h=2', 'Niveau XGB h=3'])
    for idx, (nom, d) in enumerate(sorted_d, 2):
        ws_alerts.cell(row=idx, column=1, value=d['district_court'])
        for col_offset, (algo, h) in enumerate([('RF',1),('RF',2),('RF',3),('XGB',1),('XGB',2),('XGB',3)], 2):
            niv = d.get(f"{algo}_{h}_niveau", 'vert')
            c = ws_alerts.cell(row=idx, column=col_offset, value=niv.upper())
            c.fill = alert_fill.get(niv, alert_fill['vert'])
            c.alignment = Alignment(horizontal='center')
    _autosize(ws_alerts, [22, 14, 14, 14, 14, 14, 14])
    ws_alerts.freeze_panes = 'A2'

    # ---------- Feuille : Performances ----------
    ws_perf = wb.create_sheet("Performances ML")
    _apply_header(ws_perf, ['Algorithme', 'Pli (Fold)', 'Horizon', 'RMSE', 'MAE', 'R²', 'N prédictions'])
    perfs = Performance.objects.order_by('algorithme', 'fold', 'horizon')
    for idx, p in enumerate(perfs, 2):
        ws_perf.cell(row=idx, column=1, value=p.algorithme)
        ws_perf.cell(row=idx, column=2, value=p.fold)
        ws_perf.cell(row=idx, column=3, value=p.horizon)
        ws_perf.cell(row=idx, column=4, value=round(p.rmse, 3))
        ws_perf.cell(row=idx, column=5, value=round(p.mae, 3))
        ws_perf.cell(row=idx, column=6, value=round(p.r2, 3))
        ws_perf.cell(row=idx, column=7, value=p.n_predictions)
    _autosize(ws_perf, [12, 10, 10, 10, 10, 10, 14])
    ws_perf.freeze_panes = 'A2'

    # ============================================================
    # Sauvegarde + retour HTTP
    # ============================================================
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"previsions_globales_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def api_export_excel(request):
    """Export Excel des prévisions pour le district sélectionné."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    user = request.user
    district_id = request.GET.get('district_id')
    algo = request.GET.get('algo', 'XGB').upper()
    horizon = int(request.GET.get('horizon', 3))

    district = get_object_or_404(District, pk=district_id)
    if not user.peut_voir_district(district):
        return JsonResponse({'error': 'accès refusé'}, status=403)

    # Reconstruire les prévisions (même logique que api_prevision)
    response = api_prevision(request)
    data = json.loads(response.content)
    previsions = data.get('previsions', [])
    if isinstance(previsions, dict):
        previsions = previsions.get('XGB', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prévisions"

    header_fill = PatternFill(start_color="0E476E", end_color="0E476E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Horizon (mois)", "Date cible", "Incidence prédite (/1000)",
               "Cas attendus", "Seuil P75", "Seuil P90", "Niveau d'alerte"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, p in enumerate(previsions, 2):
        ws.cell(row=row_idx, column=1, value=p['horizon'])
        ws.cell(row=row_idx, column=2, value=p.get('mois_label', p.get('date')))
        ws.cell(row=row_idx, column=3, value=p['incidence'])
        ws.cell(row=row_idx, column=4, value=p['cas'])
        ws.cell(row=row_idx, column=5, value=p['p75'])
        ws.cell(row=row_idx, column=6, value=p['p90'])
        ws.cell(row=row_idx, column=7, value=p['niveau'].upper())

    for col_letter in 'ABCDEFG':
        ws.column_dimensions[col_letter].width = 22

    buf = openpyxl.writer.excel.save_virtual_workbook(wb) if hasattr(openpyxl.writer.excel, 'save_virtual_workbook') else None
    if buf is None:
        import io
        buf_io = io.BytesIO()
        wb.save(buf_io)
        buf = buf_io.getvalue()

    filename = f"previsions_{district.nom_court}_{algo}_h{horizon}_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
